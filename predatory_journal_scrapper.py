import os
import json
from pathlib import Path
import requests as rs
import pandas as pd
import numpy as np
from openpyxl import load_workbook

from config import BASE_DIR, META_DIR, DATE_FORMAT, UTC_NOW, OBJ_LIST


def log_file_date(obj_type: str, meta_dir: str) -> None:
    """
    Create or update a JSON file with the latest processing date for a given object type.
    
    Args:
        obj_type (str): The type of object being processed (e.g., 'Journal', 'Publisher')
        base_dir (str): Base directory for file operations
    
    Returns:
        None
    """
    metadata_dir = Path(meta_dir)
    
    date_file = metadata_dir / f"{obj_type.lower()}s_latest_date.json"
    utc_now_str = UTC_NOW.strftime(DATE_FORMAT)
    
    metadata = {
        "name": obj_type,
        "latest_date": utc_now_str
    }
    
    with open(date_file, "w") as outfile:
        json.dump(metadata, outfile, indent=4)
    
    print(f"Updated datetime for {obj_type} in file {date_file} to {utc_now_str}")


def combine_dataframes(obj_type: str, df_old: pd.DataFrame, df_new: pd.DataFrame) -> pd.DataFrame:
    """
    Combine old and new dataframes, preserving historical 'Since' dates and updating links.
    
    Args:
        obj_type (str): The column name to join on. Either 'Journal' or 'Publisher'.
        df_old (pd.DataFrame): Existing dataframe with historical data
        df_new (pd.DataFrame): New dataframe with updated data
    
    Returns:
        pd.DataFrame: Combined dataframe with appropriate date and link handling
    """
    # Merge dataframes on the object type column
    df_combined = df_new.merge(
        df_old, 
        on=obj_type, 
        suffixes=('_new', '_old'), 
        how='left'
    )
    
    # Ensure both datetime columns are timezone-aware UTC for comparison
    df_combined['Since_old'] = pd.to_datetime(df_combined['Since_old'], utc=True)
    df_combined['Since_new'] = pd.to_datetime(df_combined['Since_new'], utc=True)
    
    # Keep the earliest 'Since' date (df_old)
    # Else, use value from df_new
    df_combined['Since'] = np.where(
        pd.isna(df_combined['Since_old']) | (df_combined['Since_new'] <= df_combined['Since_old']), 
        df_combined['Since_new'], 
        df_combined['Since_old']
    )
    
    # If the link are different between versions, use the new one. 
    # Check for nan Link in df_old
    df_combined['Link'] = np.where(
        pd.isna(df_combined['Link_old']) | (df_combined['Link_new'] != df_combined['Link_old']), 
        df_combined['Link_new'], 
        df_combined['Link_old']
    )
    
    return df_combined


def extract_dataframe_from_excel(obj_type: str, obj_url: str, temp_xlsx: Path) -> pd.DataFrame:
    """
    Download Excel file from URL and extract data with hyperlinks.
    
    Args:
        obj_type (str): The type of object being processed
        obj_url (str): URL to download the Excel file from
        temp_xlsx (Path): Path to save the temporary Excel file
    
    Returns:
        pd.DataFrame: Processed dataframe with object names and links
    
    Raises:
        requests.RequestException: If the file download fails
        Exception: If Excel processing fails
    """
    try:
        # Download the Excel file
        response = rs.get(url=obj_url)
        response.raise_for_status()
        
        with open(temp_xlsx, 'wb') as f:
            f.write(response.content)
        
        # Load workbook to extract hyperlinks
        wb = load_workbook(temp_xlsx)
        ws = wb['Sheet1']
        
        # Extract hyperlinks from column B (index 2)
        hyperlinks = []
        for row in range(1, ws.max_row + 1):
            try:
                cell = ws.cell(row=row, column=2)
                link = cell.hyperlink.target if cell.hyperlink else np.nan
                hyperlinks.append(link)
            except AttributeError:
                hyperlinks.append(np.nan)
        
        # Read Excel data into DataFrame
        df_new = pd.read_excel(temp_xlsx, header=None)
        
        # Set up columns
        df_new[obj_type] = df_new[1]  # Assuming column 1 contains the object names
        df_new['Link'] = hyperlinks[:len(df_new)]
        
        return df_new
        
    except rs.RequestException as e:
        print(f"Error downloading file from {obj_url}: {e}")
        raise
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        raise


def process_object_data(obj_type: str, obj_url: str, base_dir: str, meta_dir: str) -> None:
    """
    Process data for a specific object type by downloading, comparing with existing data,
    and updating the CSV file only when changes are detected.
    
    Args:
        obj_type (str): The type of object to process (e.g., 'Journal', 'Publisher')
        obj_url (str): URL to download the data from
        base_dir (str): Base directory for file operations
    
    Returns:
        None
    """
    directory = Path(base_dir)
    temp_xlsx = directory / f"{obj_type.lower()}_temp.xlsx"
    target_csv = directory / f"predatory_{obj_type.lower()}.csv"
    
    try:
        # Get new data
        df_new = extract_dataframe_from_excel(
            obj_type=obj_type, 
            obj_url=obj_url, 
            temp_xlsx=temp_xlsx
        )
        
        # Add current timestamp (ensure UTC timezone)
        df_new['Since'] = UTC_NOW
        
        if len(df_new) == 0:
            print(f"No data found for {obj_type}")
            return
        
        data_changed = True  # Flag to track if data has changed
        
        # Compare with existing data if CSV exists
        if target_csv.exists():
            df_old = pd.read_csv(target_csv)
            # Ensure old data is parsed as UTC timezone-aware
            df_old['Since'] = pd.to_datetime(df_old['Since'], utc=True)
            
            compare_cols = [obj_type, 'Link']
            
            # Check if data has changed
            if df_old[compare_cols].equals(df_new[compare_cols]):
                data_changed = False
                print(f"No changes detected for {obj_type}, skipping update")
                return
            else:
                print(f"Changes detected for {obj_type}, updating data")
                df_new = combine_dataframes(
                    obj_type=obj_type, 
                    df_old=df_old, 
                    df_new=df_new
                )
        
        # Only save and log if data has changed or if it's the first time
        if data_changed:
            # Save updated data
            df_new[[obj_type, 'Link', 'Since']].to_csv(
                target_csv, 
                index=False, 
                date_format=DATE_FORMAT
            )
            
            # Log the update only when there are changes
            log_file_date(obj_type, meta_dir)
            
            print(f"Successfully processed {len(df_new)} records for {obj_type}")
        
    except Exception as e:
        print(f"Error processing {obj_type}: {e}")
    finally:
        # Clean up temporary file
        if temp_xlsx.exists():
            temp_xlsx.unlink()

def main() -> None:
    """
    Main function to process all object types defined in OBJ_LIST.
    
    Returns:
        None
    """    
    print(f"Starting data processing for {len(OBJ_LIST)} object types...")
    
    for obj_type in OBJ_LIST:
        obj_url = os.environ.get(f'{obj_type.upper()}_URL')
        
        if not obj_url:
            print(f"Warning: No URL found for {obj_type.upper()}_URL")
            continue
            
        print(f"\nProcessing {obj_type}...")
        process_object_data(
            obj_type=obj_type, 
            obj_url=obj_url, 
            base_dir=BASE_DIR,
            meta_dir=META_DIR
        )
    
    print("\nData processing completed.")


if __name__ == "__main__":
    main()